import pyautogui as pt
from time import sleep
import pyperclip
import random
import openpyxl
import pickle
import re
from datetime import datetime
import clipboard


class DataFinder:
    def __init__(self, patterns):
        self.patterns = patterns
        self.compiledPatterns = []
        for pattern in self.patterns:
            compiled = re.compile(pattern)
            self.compiledPatterns.append(compiled)

    def check(self, StringData):
        StringData = str(StringData)
        result = []
        for cp in self.compiledPatterns:
            matches = cp.finditer(StringData)
            for match in matches:
                result.append(match)
        return result

# Reading data From From File
# Define the location of the file
class Logs:
    def __init__(self):
        self.FilledDataFile = "FilledData.txt"
        self.UnfilledDataFile = "UnfilledData.txt"
        self.InvalidRecordFile = "InvalidData.txt"
        self.lastEventInfo = 'LastEventInfo.txt'
        self.fileOpened = False

    def openAllFile(self, mode="ab"):
        self.fileOpened = True
        self.FDFile = open(self.FilledDataFile, mode)
        self.UFDFile = open(self.UnfilledDataFile, mode)
        self.IRFile = open(self.InvalidRecordFile, mode)

    def updateLastEvent(self, info):
        file = open(self.lastEventInfo, "wb")
        pickle.dump(info, file)
        file.close()

    def getLastEventInfo(self):
        try:
            file = open(self.lastEventInfo, "rb")
            data = pickle.load(file)
            file.close()
            return data
        except:
            return None

    def closeAllFile(self):
        self.fileOpened = False
        self.FDFile.close()
        self.UFDFile.close()
        self.IRFile.close()

    def readInvalidRecord(self):
        try:
            return pickle.load(self.IRFile)
        except EOFError:
            return 'EOF'

    def readFilledRecord(self):
        try:
            return pickle.load(self.FDFile)
        except EOFError:
            return 'EOF'

    def readUnfilledData(self):
        try:
            return pickle.load(self.UFDFile)
        except EOFError:
            return 'EOF'

    def clearAllTheLogs(self):
        self.openAllFile("wb")
        self.closeAllFile()

    def addFilledRecord(self, record):
        pickle.dump(record, self.FDFile)

    def addUnfilledRecord(self, record):
        pickle.dump(record, self.UFDFile)

    def addInvalidrecord(self, record):
        pickle.dump(record, self.IRFile)

class ExcelFileHandler:
    def __init__(self, fields={'Sr.No.': 1, 'District': 2, 'Block': 3, 'Village': 4,'Address': 5, 'Pradhan': 6, 'Mobile':7}, file="data.xlsx"):
        self.fields = fields  # {'name':1, 'Pardhan': 4}
        self.fileName = file
        self.LastRow = None

        # Fields
        self.SERIALNUMBER = 'Sr.No.'
        self.DISTRICT = 'District'
        self.BLOCK = 'Block'
        self.VILLAGE = 'Village'
        self.ADDRESS = 'Address'
        self.PARDHAN = 'Pradhan'
        self.MOBILE = 'Mobile'

        # Current Row
        self.CurrentRow = 2
        self.fileOpened = False
        self.datafilter = [DataFinder([r'[.]*']), DataFinder([r'[`]*']), DataFinder([r'[-]*']), DataFinder([r'[@]*']),
                           DataFinder([r'[(]*']), DataFinder([r'[)]*'])]

    def removeDots(self, string):
        finalString = str(string)
        for datafilter in self.datafilter:
            finder = datafilter
            newstr = ""
            pretart = 0
            matchs = finder.check(finalString)
            if len(matchs) == 0:
                continue
            for match in matchs:
                start, end = match.span()
                newstr += finalString[pretart: start]
                pretart = end
            finalString = newstr

        return finalString

    def openFile(self):
        self.fileOpened = True
        self.File = openpyxl.load_workbook(self.fileName)
        self.Sheet = self.File.active
        self.LastRow = self.Sheet.max_row

    def closeFile(self):
        self.File.close()
        self.fileOpened = False

    def setCurrentRow(self, row):
        if not self.fileOpened:
            print("Please open the excel File")
        if row <= self.LastRow:
            self.CurrentRow = row
            return True
        else:
            return False

    def moveOnNextRow(self):
        if self.CurrentRow <= self.LastRow:
            self.CurrentRow += 1
            return True
        else:
            return False

    def read(self, fields):
        data = {}
        for field in fields:
            read_data = self.Sheet.cell(row=self.CurrentRow, column=self.fields[field])
            if data is not None:
                read_data = read_data.value
                read_data = self.removeDots(read_data)
                data[field] = str(read_data)
            else:
                return None
        return data, self.CurrentRow


class AutoRunner:
    def __init__(self):
        self.imgLocater = 'assets/locator.png'
        self.imgPreviewsActive = 'assets/previews_active.png'
        self.imgNextActive = "assets/next_active.png"
        self.imgNext = "assets/next.png"
        self.imgPreviews = "assets/previews.png"
        self.imgTick = 'assets/tick.png'
        self.addBranch = 'assets/add.png'
        self.buttonSaveEntry = 'assets/saveentry.png'

        # add branch buttons
        self.buttonSaveBranch = 'assets/savebranch.png'
        self.buttonSarpanch = 'assets/sarpanch.png'
        self.buttonSarpanchSelected = "assets/sarpanchselected.png"
        self.fieldname = 'assets/namefield.png'
        self.fieldMobile = 'assets/mobilefield.png'
        self.fieldyes = 'assets/yes.png'
        self.field = 'assets/field.png'

        self.villageSelecter = 'assets/villageselect.png'
        self.branchYesSelecter = 'assets/branchYesSelecter.png'

        # exit buttons
        self.ButtonQuit = 'assets/quit.png'
        self.ButtonQuit2 = 'assets/quit2.png'

        # Locaters
        self.addBranchLocater = 'assets/addBranchLocater.png'
        self.keyboardLocater = 'assets/keyboard.png'

        # common buttons
        self.buttonCopy = 'assets/copy.png'

        # Pages
        self.pageDistrict = 'district'
        self.pageBlock = 'block'
        self.pageCluster = 'cluster'
        self.pageFramer = 'farmer'
        self.pageSave = 'saveEntry'

        # Branch pages
        self.branchRole = 'role'
        self.branchName = 'name'
        self.branchMobile = 'mobile'
        self.branchSelectSmartphone = 'smartphone'
        self.branchSave = 'saveBranch'

        self.pages = {self.pageDistrict:1, self.pageBlock:2, self.pageCluster:3,
                      self.pageFramer:4, self.pageSave: 5}
        self.branchPages = {self.branchRole:6, self.branchName:7, self.branchMobile:8,
                            self.branchSelectSmartphone:9, self.branchSave: 10}

        self.addingBranch = False
        self.currentBranchPage = 1
        self.currentPage = 4

    def selectAll(self):
        # pt.press('end')
        # pt.hotkey('shift', 'home')
        pt.hotkey('ctrl', 'a')

    def quitfromBranchFrom(self):
        run = True
        count = 1
        while run:
            if count > 7:
                if self.locateToAddBrach():
                    return True
                else:
                    return False
            position = pt.locateCenterOnScreen(self.buttonSarpanch, confidence=.9)
            if position is None:
                position = pt.locateCenterOnScreen(self.buttonSarpanchSelected, confidence=.9)
            if position is None:
                self.clickonPreviews()
                count += 1
            else:
                if self.clickonButton(self.ButtonQuit, retry=3):
                    if self.clickonButton(self.ButtonQuit2, retry=3):
                        sleep(1)
                        self.resetPages()
                        return True
                    else:
                        return False
                else:
                    return False
        return False

    def resetPages(self):
        self.currentPage = self.pages[self.pageFramer]
        self.currentBranchPage = self.branchPages[self.branchRole]
        self.addingBranch = False

    def fillBranchFrom(self, name, mobile):
        print("Now in Branch Form......")
        self.movetoLocater()
        self.removeCopyifThere()
        if not self.clickonButton(self.buttonSarpanch, retry=2, confidence=0.9):
            print("Unable to check the Sarpanch Button")
            return False
        if not self.clickonNext():
            print("Unable to click on next button")
            return False
        sleep(0.5)
        self.removeCopyifThere()
        if not self.clickonButton(self.fieldname, retry=3):
            if not self.clickonButton(self.field, retry=3):
                print("Unable to find the Name Field")
                return False
        for i in range(3):
            self.selectAll()
            # pt.hotkey('ctrl','a')
            clipboard.copy(name)
            pt.hotkey('ctrl','v')

        if not self.clickonNext():
            print("Unable to click on next button")
            return False
        sleep(0.5)
        self.removeCopyifThere()
        if not self.clickonButton(self.fieldMobile, retry=3):
            if not self.clickonButton(self.field, retry=3):
                print("Unable to find the Mobile Field")
                return False

        for i in range(3):
            self.selectAll()
            # pt.hotkey('ctrl','a')
            clipboard.copy(mobile)
            pt.hotkey('ctrl','v')
        if not self.clickonNext():
            print("Unable to click on next button")
            return False
        sleep(0.5)
        self.removeCopyifThere()
        if not self.clickonButton(self.fieldyes, confidence=0.9, grayscale=False, retry=3):
            print("Unable to Click on Yes Field")
            return False
        if not self.clickonNext():
            print("Unable to click on next button")
            return False
        sleep(0.5)
        if not self.clickonButton(self.buttonSaveBranch, confidence=0.9, retry=3):
            position = pt.locateCenterOnScreen(self.branchYesSelecter, confidence=0.9)
            if position is not None:
                self.moveFormCurrentPosition(0, 100)
                pt.click(button='left')
                if pt.locateCenterOnScreen(self.imgTick, confidence=0.9) is None:
                    print("Unable to click on Save Brach")
                    return False
            else:
                print("Unable to click on Save Brach")
                return False
        return True

    def chnageVillage(self, newVillage):
        if not self.gotoPage(self.pageCluster):
            return False

        self.moveFormCurrentPosition(30, 30)
        count = 0
        while True:
            if count >= 3:
                if not self.locateToAddBrach():
                    continue
                if not self.gotoPage(self.pageCluster):
                    continue
                self.moveFormCurrentPosition(30, 30)
                count = 0
            position = pt.locateCenterOnScreen(self.villageSelecter, confidence=.9)
            if position is not None:
                pt.moveTo(position[0], position[1]+100, duration=.05)
                while pt.locateCenterOnScreen(self.villageSelecter, confidence=.9) != None:
                    pt.click(button='left')
                break
            else:
                for i in range(20):
                    pt.scroll(-500)
                sleep(1)
                count += 1
        for i in range(3):
            self.selectAll()
            # pt.hotkey('ctrl', 'a')
            clipboard.copy(newVillage)
            pt.hotkey('ctrl', 'v')
        sleep(1)
        if not self.clickonNext():
            print("Unable to click on next")
            return False
        return True

    def isKeyboardActive(self):
        position = pt.locateCenterOnScreen(self.keyboardLocater, confidence=.9)
        if position is not None:
            return True
        return False

    def locateToAddBrach(self):
        run = True
        count = 1
        position = None
        while run:
            if count > 7:
                return False
            position = pt.locateCenterOnScreen(self.buttonSaveEntry, confidence=.9)
            if position is not None:
                if not self.clickonPreviews():
                    return False
                run = False
            else:
                self.clickonNext()
                count += 1

        count = 1
        while position is None and count <= 3:
            position = pt.locateCenterOnScreen(self.addBranch, confidence=.9)
            count += 1
        if position is not None:
            self.resetPages()
            return True
        return False

    def removeCopyifThere(self):
        if not self.clickonButton(self.buttonCopy, retry=1, grayscale=True, confidence=0.8):
            return True
        return False

    def moveFormCurrentPosition(self, pixel_x=0, pixel_y=0):
        x, y = pt.position()
        x += pixel_x
        y += pixel_y
        pt.moveTo(x, y)

    def gotoPage(self, page):
        if self.addingBranch:
            if page in self.pages:
                if not self.quitfromBranchFrom():
                    print("Unable to quit from add Branch From,Please Quit manually and in the Menu Chose option: Add Branch From manually Closed")
                    return False

        if self.addingBranch:
            if page in self.branchPages:
                pageNumber = self.branchPages[page]
                while pageNumber != self.currentBranchPage:
                    if self.currentBranchPage < pageNumber:
                        if not self.clickonNext(retry=3):
                            return False
                    if self.currentBranchPage > pageNumber:
                        if not self.clickonPreviews(retry=3):
                            return False
                return True
            else:
                return False
        else:
            if page in self.pages:
                pageNumber = self.pages[page]
                while pageNumber != self.currentPage:
                    if self.currentPage < pageNumber:
                        if not self.clickonNext(retry=3):
                            return False
                    elif self.currentPage > pageNumber:
                        if not self.clickonPreviews(retry=3):
                            return False
                return True
            else:
                return False

    def clickonButton(self, button, confidence=.9,  grayscale=False, retry=1):
        count = 1
        while count <= retry:
            position = pt.locateCenterOnScreen(button, confidence=confidence, grayscale=grayscale)
            if position is not None:
                pt.moveTo(position[0], position[1], duration=.05)
                pt.click(button='left')
                if button == self.buttonSaveBranch:
                    self.addingBranch = False
                    self.currentBranchPage = self.branchPages[self.branchRole]
                elif button == self.addBranch:
                    self.addingBranch = True
                    self.currentBranchPage = self.branchPages[self.branchRole]
                return True
            count += 1
        return False

    def clickOnAddBranch(self):
        self.removeCopyifThere()
        if self.clickonButton(self.addBranch, confidence=.9, retry=3):
            return True
        return False
        # position = pt.locateCenterOnScreen(self.addBranch, confidence=.9)
        # if position is not None:
        #     self.addingBranch = True
        #     pt.moveTo(position[0], position[1], duration=.05)
        #     pt.click(button='left')
        #     return True
        # return False

    def movetoLocater(self):
        position = pt.locateCenterOnScreen('assets/locator.png', confidence=.9)
        if position is not None:
            pt.moveTo(position[0], position[1], duration=.05)
            return True
        return False

    def clickonNext(self, retry=1):
        position = pt.locateCenterOnScreen(self.imgNextActive, confidence=.9)
        if position is not None:
            x, y = position
            pt.moveTo(x, y, duration=.05)
            pt.click(button='left')
            if self.addingBranch:
                self.currentBranchPage += 1
            else:
                self.currentPage += 1
            return True
        else:
            return False

    def clickonPreviews(self, retry=1):
        count = 1
        while count <= retry:
            position = pt.locateCenterOnScreen(self.imgPreviewsActive, confidence=.9)
            if position is not None:
                x, y = position
                pt.moveTo(x, y, duration=.05)
                pt.click(button='left')
                if self.addingBranch:
                    self.currentBranchPage -= 1
                else:
                    self.currentPage -= 1
                return True
            count += 1
        return False

    def isDeactive_Next(self):
        position = pt.locateCenterOnScreen(self.imgNext, confidence=.9)
        if position is not None:
            return True
        return False

    def isDeactive_Previews(self):
        position = pt.locateCenterOnScreen(self.imgPreviews, confidence=.9)
        if position is not None:
            return True
        return False

def viewLogs():
    logs = Logs()
    excel = ExcelFileHandler(file='data.xlsx')
    run = True
    while run:
        print("-------- Logs Meuu ---------")
        print("1. Filled Record")
        print("2. Unfilled Record")
        print("3. Invalid record")
        print("4. Get Last Event Info")
        print("5. Clear All Logs")
        print("6. Go Back to Main Menu")
        choice = input("Enter the choice: ")
        if choice == '1':
            reading = True
            logs.openAllFile('rb')
            count = 0
            while reading:
                data = logs.readFilledRecord()
                if data == 'EOF':
                    reading = False
                else:
                    count += 1
                    print(data)
            print(f"Total Filled Form: {count}")
            logs.closeAllFile()
        elif choice == '2':
            reading = True
            logs.openAllFile('rb')
            while reading:
                data = logs.readUnfilledData()
                if data == 'EOF':
                    reading = False
                else:
                    print(data)
            logs.closeAllFile()
        elif choice == '3':
            reading = True
            logs.openAllFile('rb')
            while reading:
                data = logs.readInvalidRecord()
                if data == 'EOF':
                    reading = False
                else:
                    print(data)
            logs.closeAllFile()
        elif choice == '4':
            data = logs.getLastEventInfo()
            print(data)
        elif choice == '5':
            logs.clearAllTheLogs()
        elif choice == '6':
            return
        else:
            print("Invalid Choice")

def Filldata():
    logs = Logs()
    excel = ExcelFileHandler(file='data.xlsx')
    invalidDatafinder = DataFinder([r'[?]*[?]', 'None'])
    autoRunner = AutoRunner()
    run = True
    while run:
        print("------- Auto Fill Menu -------")
        print("Informations_________________________________________________")
        print(f"Logs Files Open State: {logs.fileOpened}, Excel File Open State: {excel.fileOpened}")
        print(f"Corrent Reading Row: {excel.CurrentRow}")
        print("Options____________________________")
        print("1. Set Current Row From were you Want to Fill")
        print("2. Open Excel File")
        print("3. Reset the Excel File")
        print("4. Start Filling")
        print("5. Close Excel File")
        print("6. Add Branch From manually Closed")
        print("7. Goto the main Menu")
        print("8: Count All Invalid Record in excel Sheet")
        choice = input("Enter teh Choice : ")
        if choice == '1':
            row = int(input("Enter the Row: "))
            data = excel.setCurrentRow(row)
            if not data:
                print("Invalid Row, Please Try Again")
        elif choice == '2':
            print("Opening the excel File..... PLease Wait")
            try:
                excel.openFile()
            except:
                excel.fileOpened = False
                print("Error in opening the excel File, Please check the system..")
                continue
            print("Excel File Open successfully")
        elif choice == '3':
            try:
                excel.closeFile()
            except:
                pass
            excel = ExcelFileHandler(file='data.xlsx')
        elif choice == '4':
            print("--------------------------------- Now Starting The Filling Process ----------------------------")
            autoOpenExcelFile = False
            if not excel.fileOpened:
                autoOpenExcelFile = True
                print("Opening the excel File..... PLease Wait")
                try:
                    excel.openFile()
                except:
                    print("Error in opening the excel File, Please check the system..")
                    continue
                print("Excel File Open successfully")
            choice = input("Do you want to start from spacific Row (yes/no): ")
            if choice == 'yes':
                print(f"Last Row if Sheet is: {excel.LastRow}")
                row = int(input("Enter the Row Number: "))
                if not excel.setCurrentRow(row):
                    print("Unable to set Row PLase Give Valid Row Number or Try Again..")
                    continue
            print("Now Opening the Log files")
            try:
                logs.openAllFile()
            except:
                print("Unable to open Log Files")
                continue
            print("Log Files Successfully open")
            totalRecord = int(input("Enter How Many record You want to Fill: "))
            startPoint = input("From were you want to start (1: From 1st Row, 2: From Current Row, 3: From Last Record): ")
            if startPoint == '1':
                excel.setCurrentRow(2)
            elif startPoint == '2':
                print("Current Row is:", excel.CurrentRow)
            elif startPoint == '3':
                try:
                    lastRecord = logs.getLastEventInfo()
                except:
                    print('Unable  to fetch last record :(')
                    continue
                if lastRecord is None:
                    print('Unable  to fetch last record :(')
                    continue
                else:
                    lastRow = lastRecord['lastRecord']['row']
                    excel.setCurrentRow(int(lastRow)+1)

            startFlag = True

            count = 1
            StartFilling = True
            PreviusRecord = logs.getLastEventInfo()
            if PreviusRecord is not None:
                PreviusRecord = PreviusRecord['lastRecord']

            while StartFilling and count <= totalRecord:
                data, row = excel.read([excel.SERIALNUMBER, excel.DISTRICT, excel.BLOCK,
                                        excel.VILLAGE, excel.PARDHAN, excel.MOBILE])
                print(f"Page Info: Record Number: {count} Main Page: {autoRunner.currentPage}, Branch Page: {autoRunner.currentBranchPage}, Branch Page State: {autoRunner.addingBranch}")

                if startFlag:
                    value = pt.confirm(text=str(data), title='Starting Record', buttons=['OK', 'CANCEL', '1 Record Back'])
                    if value == 'CANCEL' or value == 'cancel':
                        StartFilling = False
                        break
                    elif value == '1 Record Back':
                        if row > 2:
                            if excel.setCurrentRow(row-1):
                                continue
                            else:
                                print("Unable to move Back")
                                StartFilling = False
                                break
                    startFlag = False

                # print(data)
                InvalidDataFlag = False
                for key, value in data.items():
                    invalidData = invalidDatafinder.check(value)
                    if len(invalidData) > 0:
                        InvalidDataFlag = True
                        data['row'] = str(row)
                        logs.addInvalidrecord(data)
                        break

                if InvalidDataFlag:
                    if not excel.moveOnNextRow():
                        print("Unable to Move On Next Row, May Be it was Last Row....")
                        StartFilling = False
                    print(f"Invalid data: {data}")
                    continue

                FormFilled = False
                # Here Automation--------------------------------------------
                while True:
                    # print(PreviusRecord)
                    # Selecting District
                    if PreviusRecord is not None:
                        if PreviusRecord[excel.DISTRICT] != data[excel.DISTRICT]:
                            if not autoRunner.gotoPage(autoRunner.pageDistrict):
                                print("Unable to Locate to the District Page")
                                break
                            else:
                                pt.confirm(text=f'District is: {data[excel.DISTRICT]}\nPlease select the District then Press Ok', title='Chnage District', buttons=['OK'])
                    else:
                        if not autoRunner.gotoPage(autoRunner.pageDistrict):
                            print("Unable to Locate to the District Page")
                            break
                        else:
                            pt.confirm(text=f'District is: {data[excel.DISTRICT]}\nPlease select the District then Press Ok', title='Chnage District', buttons=['OK'])

                    # Selecting Block
                    if PreviusRecord is not None:
                        if PreviusRecord[excel.BLOCK] != data[excel.BLOCK]:
                            if not autoRunner.gotoPage(autoRunner.pageBlock):
                                print("Unable to Locate to the Block Page")
                                break
                            else:
                                pt.confirm(text=f'Block is: {data[excel.BLOCK]}\nPlease select the Block then Press Ok', title='Chnage Block', buttons=['OK'])
                    else:
                        if not autoRunner.gotoPage(autoRunner.pageBlock):
                            print("Unable to Locate to the Block Page")
                            break
                        else:
                            pt.confirm(text=f'Block is: {data[excel.BLOCK]}\nPlease select the Block then Press Ok', title='Chnage Block', buttons=['OK'])

                    if PreviusRecord is not None:
                        if PreviusRecord[excel.VILLAGE] != data[excel.VILLAGE]:
                            if not autoRunner.chnageVillage(data[excel.VILLAGE]):
                                break
                    else:
                        if not autoRunner.chnageVillage(data[excel.VILLAGE]):
                            break

                    sleep(2)
                    if not autoRunner.clickOnAddBranch():
                        break
                    if not autoRunner.fillBranchFrom(data[excel.PARDHAN], data[excel.MOBILE]):
                        if autoRunner.addingBranch:
                            if autoRunner.quitfromBranchFrom():
                                continue
                        break

                    sleep(1)
                    FormFilled = True
                    break

                # Automation End---------------------------------------------

                if not FormFilled:
                    print("Unfilled Form: ", data)
                    data['row'] = str(row)
                    logs.addUnfilledRecord(data)
                    StartFilling = False
                    continue

                # saving log of filled Form
                print("filled Form: ", data)
                data['row'] = str(row)
                logs.addFilledRecord(data)

                # Saving Data of Last Event..
                now = datetime.now()
                dateTime = now.strftime("%d/%m/%Y %H:%M:%S")  # dd/mm/YY H:M:S
                data['row'] = str(row)
                eventdata = {'lastRecord': data, 'time': dateTime}
                PreviusRecord = data
                logs.updateLastEvent(eventdata)

                if not excel.moveOnNextRow():
                    print("Unable to Move On Next Row, May Be it was Last Row....")
                    StartFilling = False
                count += 1

            logs.closeAllFile()
            if autoOpenExcelFile:
                excel.closeFile()
        elif choice == '5':
            excel.closeFile()
            print("File Closed")
        elif choice == '6':
            autoRunner.resetPages()
        elif choice == '7':
            return
        elif choice == '8':
            if not excel.fileOpened:
                print("Opening the excel File..... PLease Wait")
                try:
                    excel.openFile()
                except:
                    print("Error in opening the excel File, Please check the system..")
                    continue
                print("Excel File Open successfully")
            Counter = 0
            while True:
                data, row = excel.read([excel.SERIALNUMBER, excel.DISTRICT, excel.BLOCK,
                                        excel.VILLAGE, excel.PARDHAN, excel.MOBILE])
                for key, value in data.items():
                    invalidData = invalidDatafinder.check(value)
                    if len(invalidData) > 0:
                        Counter += 1
                        break
                if not excel.moveOnNextRow():
                    break
            print("Total Invalid Record is :",Counter)



run = True
while run:
    print("Warning: Don't Exit Direct Please Use the Exit Option, Otherwise Logs will get delete :(")
    print("---------- Automating Feeding -------")
    print("1. View Logs")
    print("2. Start Filling")
    print("3. Exit")
    choice = input('Enter the Choice: ')
    if choice == '1':
        viewLogs()
    elif choice == '2':
        Filldata()
    elif choice == '3':
        run = False
